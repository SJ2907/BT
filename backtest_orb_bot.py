"""
Intraday ORB + VWAP Strategy BACKTEST (Indian Stock Market - NSE)
=====================================================================

WHAT THIS DOES
--------------
Replays the SAME signal logic as the live bot (intraday_orb_bot.py) --
Opening Range Breakout + VWAP + volume + EMA trend + RSI + MACD, with
stop-loss from real swing structure and target from real pivot-point
resistance/support -- against HISTORICAL data, so you can see how the
strategy would have actually performed before ever running it live.

For every trading day, for every stock in today's Nifty 100 list, it:
  1. Builds the day's indicators from that day's actual 5-minute candles
  2. Looks for the first candle where a real signal would have fired
  3. Walks forward through the REST of that day's candles to see whether
     the stop or target got touched first (using High/Low, not just Close,
     for a more realistic "did price actually touch this level" check)
  4. If neither is touched by end of day, squares off at the last close
     (matching how the live bot forces an end-of-day square-off)

Capital compounds day-by-day: each day's starting capital is used to size
that day's position (10% risk per trade, by default), and the day's total
P&L is added before moving to the next day.

Results go into backtest_journal.xlsx: every simulated trade, plus a
Summary sheet (win rate, total P&L, return %, max drawdown) and a
BySymbol sheet (per-stock performance) -- so you can judge the setup's
actual historical accuracy before trusting it with real capital.

============================================================================
IMPORTANT LIMITATIONS -- READ BEFORE TRUSTING THE OUTPUT
============================================================================
- Yahoo Finance only provides 5-minute candles for the last ~60 CALENDAR
  DAYS (roughly 40-42 trading days). This is a hard limit on their end --
  requesting a longer period for 5m data will just silently return less
  than you asked for. This script uses the longest window Yahoo allows.
- Today's Nifty 100 list is used for the whole backtest window. The ACTUAL
  index composition a few months ago may have been slightly different
  (stocks get added/removed periodically) -- this is a small survivorship-
  bias source, not something this script corrects for.
- No brokerage, STT, slippage, or bid-ask spread is modeled. Real trading
  costs and imperfect fills will make live results worse than this backtest
  shows, typically by a noticeable margin for high-frequency strategies.
- One trade per stock per day, sized off that day's starting capital --
  this does NOT model true simultaneous portfolio risk if many stocks
  signal on the same day (i.e. it assumes your whole capital could have
  been available to every signal that day, which overstates how much you
  could really deploy if 15 stocks all fired at once).
- A ~2-3 month sample is still fairly small. Use this to sanity-check
  whether the strategy looks reasonable or obviously broken -- not as a
  precise, statistically confident estimate of long-run performance.
- THIS IS A SIMULATION FOR RESEARCH PURPOSES. It is not a guarantee of
  future results, live or otherwise.

REQUIREMENTS
------------
    pip install yfinance pandas pytz openpyxl requests

HOW TO RUN
----------
    python3 backtest_orb_bot.py
    python3 backtest_orb_bot.py --tickers RELIANCE.NS,TCS.NS   # just a few, for a quick check
    python3 backtest_orb_bot.py --days 40                       # override lookback (max ~60)

This can take a while (fetching + simulating ~100 stocks x ~40 days) --
progress is printed to the console as it goes.
"""

import argparse
import json
import os
import sys
import threading
import time
from datetime import datetime, timedelta
from http.server import BaseHTTPRequestHandler, HTTPServer
from io import StringIO
from urllib.parse import urlparse, parse_qs

import pandas as pd
import pytz
import requests
import yfinance as yf
from openpyxl import Workbook

# ============================== CONFIG ===================================

BACKTEST_CAPITAL = 300000.0     # starting capital for the simulation
RISK_PER_TRADE_PCT = 10.0        # % of capital risked per trade (as requested)
MIN_RR_RATIO = 1.5                # skip a setup if real structure doesn't offer this

OPENING_RANGE_MINUTES = 15
CANDLE_INTERVAL = "5m"
MAX_LOOKBACK_DAYS = 59            # Yahoo's real ceiling for 5m data is ~60 days

VOLUME_MULTIPLIER = 1.5
EMA_FAST, EMA_SLOW = 9, 21
RSI_PERIOD = 14
RSI_BULLISH_MIN, RSI_BULLISH_MAX = 50, 70
RSI_BEARISH_MIN, RSI_BEARISH_MAX = 30, 50
MACD_FAST, MACD_SLOW, MACD_SIGNAL = 12, 26, 9
SWING_LOOKBACK_CANDLES = 12

IST = pytz.timezone("Asia/Kolkata")
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "backtest_journal.xlsx")
NIFTY100_SOURCE_URL = "https://niftyindices.com/IndexConstituent/ind_nifty100list.csv"

FALLBACK_TICKERS = [
    "RELIANCE.NS", "TCS.NS", "HDFCBANK.NS", "ICICIBANK.NS", "INFY.NS",
    "HINDUNILVR.NS", "ITC.NS", "SBIN.NS", "BHARTIARTL.NS", "KOTAKBANK.NS",
    "LT.NS", "BAJFINANCE.NS", "HCLTECH.NS", "AXISBANK.NS", "ASIANPAINT.NS",
    "MARUTI.NS", "SUNPHARMA.NS", "TITAN.NS", "ULTRACEMCO.NS", "WIPRO.NS",
]

# ============================ END CONFIG ==================================


def get_nifty_100_tickers():
    """Fetches TODAY's official Nifty 100 list (see limitation notes above
    about using today's list for a historical window)."""
    try:
        resp = requests.get(NIFTY100_SOURCE_URL,
                             headers={"User-Agent": "Mozilla/5.0"}, timeout=20)
        resp.raise_for_status()
        df = pd.read_csv(StringIO(resp.text))
        symbol_col = "Symbol" if "Symbol" in df.columns else df.columns[2]
        tickers = [f"{s.strip()}.NS" for s in df[symbol_col].dropna().tolist()]
        if len(tickers) < 90:
            raise ValueError(f"Only got {len(tickers)} tickers, expected ~100")
        print(f"Fetched live Nifty 100 list: {len(tickers)} tickers.")
        return tickers
    except Exception as e:
        print(f"[warn] Live Nifty 100 fetch failed ({e}). Using a small fallback list instead.")
        return FALLBACK_TICKERS


# ------------------------------ INDICATORS --------------------------------
# Identical math to the live bot -- all are causal (each row's value only
# depends on that row and earlier ones), so computing them once over a full
# day's candles gives the exact same result as computing them incrementally
# candle-by-candle. No look-ahead bias from doing it this way.

def compute_vwap(df):
    typical_price = (df["High"] + df["Low"] + df["Close"]) / 3
    df["cum_vol"] = df["Volume"].cumsum()
    df["cum_vol_price"] = (typical_price * df["Volume"]).cumsum()
    df["VWAP"] = df["cum_vol_price"] / df["cum_vol"]
    return df


def compute_ema(df, span, col_name):
    df[col_name] = df["Close"].ewm(span=span, adjust=False).mean()
    return df


def compute_rsi(df, period, col_name="RSI"):
    delta = df["Close"].diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.ewm(alpha=1 / period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1 / period, adjust=False).mean()
    rs = avg_gain / avg_loss.replace(0, pd.NA)
    df[col_name] = 100 - (100 / (1 + rs))
    df[col_name] = df[col_name].fillna(50)
    return df


def compute_macd(df, fast, slow, signal):
    ema_fast = df["Close"].ewm(span=fast, adjust=False).mean()
    ema_slow = df["Close"].ewm(span=slow, adjust=False).mean()
    df["MACD"] = ema_fast - ema_slow
    df["MACD_SIGNAL"] = df["MACD"].ewm(span=signal, adjust=False).mean()
    return df


def get_opening_range(df):
    if df.empty:
        return None, None
    session_start = df.index[0]
    cutoff = session_start + pd.Timedelta(minutes=OPENING_RANGE_MINUTES)
    opening_slice = df[df.index <= cutoff]
    if opening_slice.empty:
        return None, None
    return opening_slice["High"].max(), opening_slice["Low"].min()


def classic_pivots(H, L, C):
    P = (H + L + C) / 3
    return {
        "P": P, "R1": 2 * P - L, "S1": 2 * P - H,
        "R2": P + (H - L), "S2": P - (H - L),
        "R3": H + 2 * (P - L), "S3": L - 2 * (H - P),
    }


# ------------------------------ DATA FETCHING ------------------------------

def fetch_intraday_history(ticker, days):
    try:
        df = yf.download(ticker, period=f"{days}d", interval=CANDLE_INTERVAL,
                          progress=False, auto_adjust=False)
    except Exception as e:
        print(f"  [warn] intraday fetch failed for {ticker}: {e}")
        return None
    if df is None or df.empty:
        return None
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    return df


def fetch_daily_history(ticker, days):
    # fetch extra daily history so every day in the intraday window has a
    # genuine PRIOR trading day available for pivot calculation
    try:
        df = yf.download(ticker, period=f"{days + 15}d", interval="1d",
                          progress=False, auto_adjust=False)
    except Exception as e:
        print(f"  [warn] daily fetch failed for {ticker}: {e}")
        return None
    if df is None or df.empty:
        return None
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    return df


# ------------------------------ SIMULATION CORE ----------------------------

def simulate_ticker_day(day_df, pivots, capital_start_of_day):
    """Runs the ORB+VWAP+EMA+RSI+MACD logic on one ticker's one day of
    candles. Returns a trade dict if a signal fired that day, else None."""
    if len(day_df) < max(EMA_SLOW, MACD_SLOW, RSI_PERIOD) // 3 + 4:
        return None

    day_df = compute_vwap(day_df)
    day_df = compute_ema(day_df, EMA_FAST, "EMA_FAST")
    day_df = compute_ema(day_df, EMA_SLOW, "EMA_SLOW")
    day_df = compute_rsi(day_df, RSI_PERIOD)
    day_df = compute_macd(day_df, MACD_FAST, MACD_SLOW, MACD_SIGNAL)

    or_high, or_low = get_opening_range(day_df)
    if or_high is None:
        return None

    session_start = day_df.index[0]
    or_cutoff = session_start + pd.Timedelta(minutes=OPENING_RANGE_MINUTES)

    avg_volume_full_day = day_df["Volume"].mean()
    if pd.isna(avg_volume_full_day) or avg_volume_full_day == 0:
        return None

    entry_idx = None
    signal = None

    for i in range(len(day_df)):
        row = day_df.iloc[i]
        if row.name <= or_cutoff:
            continue  # still inside the opening range window itself

        avg_volume = day_df["Volume"].iloc[:i].mean() if i > 0 else avg_volume_full_day
        if pd.isna(avg_volume) or avg_volume == 0:
            continue
        volume_ok = row["Volume"] >= VOLUME_MULTIPLIER * avg_volume

        bull_trend_ok = row["EMA_FAST"] > row["EMA_SLOW"]
        bear_trend_ok = row["EMA_FAST"] < row["EMA_SLOW"]
        bull_rsi_ok = RSI_BULLISH_MIN <= row["RSI"] <= RSI_BULLISH_MAX
        bear_rsi_ok = RSI_BEARISH_MIN <= row["RSI"] <= RSI_BEARISH_MAX
        bull_macd_ok = row["MACD"] > row["MACD_SIGNAL"]
        bear_macd_ok = row["MACD"] < row["MACD_SIGNAL"]

        if (row["Close"] > or_high and row["Close"] > row["VWAP"] and volume_ok
                and bull_trend_ok and bull_rsi_ok and bull_macd_ok):
            entry_idx, signal = i, "BULLISH"
            break
        elif (row["Close"] < or_low and row["Close"] < row["VWAP"] and volume_ok
                and bear_trend_ok and bear_rsi_ok and bear_macd_ok):
            entry_idx, signal = i, "BEARISH"
            break

    if entry_idx is None:
        return None

    entry_row = day_df.iloc[entry_idx]
    price = float(entry_row["Close"])

    swing_window = day_df.iloc[max(0, entry_idx - SWING_LOOKBACK_CANDLES):entry_idx]
    if swing_window.empty:
        return None
    swing_high, swing_low = float(swing_window["High"].max()), float(swing_window["Low"].min())

    if signal == "BULLISH":
        stop = max(or_low, swing_low)
        if stop >= price:
            return None
        candidates = [pivots[k] for k in ("R1", "R2", "R3") if pivots[k] > price]
        if not candidates:
            return None
        target = min(candidates)
    else:
        stop = min(or_high, swing_high)
        if stop <= price:
            return None
        candidates = [pivots[k] for k in ("S1", "S2", "S3") if pivots[k] < price]
        if not candidates:
            return None
        target = max(candidates)

    risk = abs(price - stop)
    reward = abs(target - price)
    if risk <= 0:
        return None
    rr_ratio = reward / risk
    if rr_ratio < MIN_RR_RATIO:
        return None

    risk_amount = capital_start_of_day * (RISK_PER_TRADE_PCT / 100)
    qty = max(1, int(risk_amount // risk))

    # Walk forward through the REST of the day to see what actually happens
    outcome, exit_price, exit_time = None, None, None
    for j in range(entry_idx + 1, len(day_df)):
        future_row = day_df.iloc[j]
        if signal == "BULLISH":
            if future_row["High"] >= target:
                outcome, exit_price = "WIN", target
            elif future_row["Low"] <= stop:
                outcome, exit_price = "LOSS", stop
        else:
            if future_row["Low"] <= target:
                outcome, exit_price = "WIN", target
            elif future_row["High"] >= stop:
                outcome, exit_price = "LOSS", stop
        if outcome:
            exit_time = future_row.name
            break

    if outcome is None:
        # never hit either level -- square off at the last candle of the day
        last_row = day_df.iloc[-1]
        exit_price = float(last_row["Close"])
        exit_time = last_row.name
        if signal == "BULLISH":
            outcome = "SQUARED_OFF_WIN" if exit_price > price else "SQUARED_OFF_LOSS"
        else:
            outcome = "SQUARED_OFF_WIN" if exit_price < price else "SQUARED_OFF_LOSS"

    if signal == "BULLISH":
        pnl = (exit_price - price) * qty
    else:
        pnl = (price - exit_price) * qty
    pnl_pct = round((pnl / (price * qty)) * 100, 2) if price * qty else 0

    return {
        "EntryDate": entry_row.name.strftime("%Y-%m-%d"),
        "EntryTime": entry_row.name.strftime("%H:%M:%S"),
        "Ticker": None,  # filled in by caller
        "Signal": signal, "Entry": round(price, 2), "Stop": round(stop, 2),
        "Target": round(target, 2), "RRRatio": round(rr_ratio, 2), "Qty": qty,
        "RiskAmount": round(risk_amount, 2),
        "ExitTime": exit_time.strftime("%H:%M:%S"), "ExitPrice": round(exit_price, 2),
        "Outcome": outcome, "PnL": round(pnl, 2), "PnLPct": pnl_pct,
    }


def run_backtest(tickers, days):
    print(f"Fetching {days}-day intraday history for {len(tickers)} tickers "
          f"(this can take a while)...")

    intraday_data = {}
    daily_data = {}
    for i, ticker in enumerate(tickers, 1):
        print(f"  [{i}/{len(tickers)}] {ticker}...")
        idf = fetch_intraday_history(ticker, days)
        ddf = fetch_daily_history(ticker, days)
        if idf is not None and ddf is not None:
            intraday_data[ticker] = idf
            daily_data[ticker] = ddf

    print(f"\nGot usable data for {len(intraday_data)}/{len(tickers)} tickers. Simulating...")

    all_days = sorted(set(
        d.date() for df in intraday_data.values() for d in df.index
    ))

    capital = BACKTEST_CAPITAL
    all_trades = []
    capital_curve = [capital]

    for day in all_days:
        day_trades = []
        for ticker, idf in intraday_data.items():
            day_df = idf[idf.index.date == day].copy()
            if day_df.empty:
                continue

            ddf = daily_data.get(ticker)
            if ddf is None:
                continue
            prior_days = ddf[ddf.index.date < day]
            if prior_days.empty:
                continue
            prev = prior_days.iloc[-1]
            pivots = classic_pivots(float(prev["High"]), float(prev["Low"]), float(prev["Close"]))

            try:
                trade = simulate_ticker_day(day_df, pivots, capital)
            except Exception as e:
                print(f"  [warn] simulation error for {ticker} on {day}: {e}")
                continue

            if trade:
                trade["Ticker"] = ticker
                day_trades.append(trade)

        if day_trades:
            day_pnl = sum(t["PnL"] for t in day_trades)
            capital += day_pnl
            for t in day_trades:
                t["CapitalAfter"] = round(capital, 2)
            all_trades.extend(day_trades)
            capital_curve.append(capital)
            print(f"  {day}: {len(day_trades)} trade(s), day P&L Rs.{round(day_pnl,2)}, "
                  f"capital now Rs.{round(capital,2)}")

    return all_trades, capital, capital_curve


# ------------------------------ EXCEL OUTPUT --------------------------------

TRADE_COLUMNS = ["EntryDate", "EntryTime", "Ticker", "Signal", "Entry", "Stop", "Target",
                  "RRRatio", "Qty", "RiskAmount", "ExitTime", "ExitPrice", "Outcome",
                  "PnL", "PnLPct", "CapitalAfter"]


def max_drawdown_pct(curve):
    peak = curve[0]
    max_dd = 0.0
    for value in curve:
        peak = max(peak, value)
        dd = (peak - value) / peak * 100 if peak else 0
        max_dd = max(max_dd, dd)
    return round(max_dd, 2)


def write_excel(trades, starting_capital, ending_capital, capital_curve):
    wb = Workbook()

    ws = wb.active
    ws.title = "Trades"
    ws.append(TRADE_COLUMNS)
    for t in trades:
        ws.append([t.get(c, "") for c in TRADE_COLUMNS])

    wins = sum(1 for t in trades if "WIN" in t["Outcome"])
    losses = sum(1 for t in trades if "LOSS" in t["Outcome"])
    total = len(trades)
    win_rate = round(wins / total * 100, 2) if total else 0
    total_pnl = round(sum(t["PnL"] for t in trades), 2)
    return_pct = round((ending_capital - starting_capital) / starting_capital * 100, 2)
    avg_rr = round(sum(t["RRRatio"] for t in trades) / total, 2) if total else 0

    summary_ws = wb.create_sheet("Summary")
    summary_rows = [
        ("Starting Capital", starting_capital),
        ("Ending Capital", round(ending_capital, 2)),
        ("Total Return %", return_pct),
        ("Total Trades", total),
        ("Wins", wins),
        ("Losses", losses),
        ("Win Rate %", win_rate),
        ("Total P&L", total_pnl),
        ("Average R:R Achieved", avg_rr),
        ("Max Drawdown %", max_drawdown_pct(capital_curve)),
        ("Risk Per Trade %", RISK_PER_TRADE_PCT),
        ("Min R:R Filter", MIN_RR_RATIO),
    ]
    for row in summary_rows:
        summary_ws.append(row)

    by_symbol = {}
    for t in trades:
        s = by_symbol.setdefault(t["Ticker"], {"trades": 0, "wins": 0, "pnl": 0.0})
        s["trades"] += 1
        s["wins"] += 1 if "WIN" in t["Outcome"] else 0
        s["pnl"] += t["PnL"]

    symbol_ws = wb.create_sheet("BySymbol")
    symbol_ws.append(["Ticker", "Trades", "Wins", "WinRate%", "TotalPnL"])
    for ticker, s in sorted(by_symbol.items(), key=lambda x: -x[1]["pnl"]):
        wr = round(s["wins"] / s["trades"] * 100, 2) if s["trades"] else 0
        symbol_ws.append([ticker, s["trades"], s["wins"], wr, round(s["pnl"], 2)])

    wb.save(OUTPUT_FILE)


# --------------------------- STANDALONE WEB SERVER --------------------------
# This lets the BACKTEST run as its own independent Render "Web Service"
# (kept fully separate from the live bot, as requested). It binds a port
# immediately (satisfying Render's requirement), then waits for you to
# trigger a run by visiting a URL -- since a backtest can take several
# minutes, running it as an immediate blocking HTTP request would time out.

backtest_status = {
    "running": False,
    "started_at": None,
    "finished_at": None,
    "result_summary": None,
    "error": None,
}


def _run_backtest_background(days, tickers_arg):
    backtest_status["running"] = True
    backtest_status["started_at"] = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    backtest_status["error"] = None
    try:
        tickers = ([t.strip() for t in tickers_arg.split(",") if t.strip()]
                   if tickers_arg else get_nifty_100_tickers())
        days = min(days, MAX_LOOKBACK_DAYS)
        trades, ending_capital, curve = run_backtest(tickers, days)
        if trades:
            write_excel(trades, BACKTEST_CAPITAL, ending_capital, curve)
            wins = sum(1 for t in trades if "WIN" in t["Outcome"])
            backtest_status["result_summary"] = (
                f"{len(trades)} trades, win rate {round(wins/len(trades)*100,1)}%, "
                f"ended at Rs.{round(ending_capital,2)} (started Rs.{BACKTEST_CAPITAL:,.2f}), "
                f"max drawdown {max_drawdown_pct(curve)}%"
            )
        else:
            backtest_status["result_summary"] = "No trades were generated -- check server logs."
    except Exception as e:
        backtest_status["error"] = str(e)
    finally:
        backtest_status["running"] = False
        backtest_status["finished_at"] = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")


class _BacktestHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        if self.path.startswith("/run-backtest"):
            self._trigger()
        elif self.path.startswith("/status"):
            self._status()
        elif self.path.startswith("/download"):
            self._download()
        else:
            self._home()

    def do_HEAD(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()

    def _home(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()
        lines = [
            "Backtest service alive.",
            "",
            "Visit /run-backtest to start a backtest (runs in the background).",
            "  Optional: /run-backtest?days=30&tickers=RELIANCE.NS,TCS.NS",
            "Visit /status to check progress.",
            "Visit /download to get backtest_journal.xlsx once finished.",
        ]
        self.wfile.write(("\n".join(lines) + "\n").encode("utf-8"))

    def _trigger(self):
        if backtest_status["running"]:
            self.send_response(200)
            self.send_header("Content-Type", "text/plain")
            self.end_headers()
            self.wfile.write(b"A backtest is already running. Check /status.\n")
            return
        query = parse_qs(urlparse(self.path).query)
        days = int(query.get("days", [MAX_LOOKBACK_DAYS])[0])
        tickers_arg = query.get("tickers", [None])[0]
        thread = threading.Thread(target=_run_backtest_background,
                                   args=(days, tickers_arg), daemon=True)
        thread.start()
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()
        self.wfile.write(
            (f"Backtest started ({days} days). This can take several minutes for the\n"
             f"full Nifty 100 list. Check /status for progress, and /download once done.\n"
             ).encode("utf-8")
        )

    def _status(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()
        lines = [
            f"Running: {backtest_status['running']}",
            f"Started at: {backtest_status['started_at']}",
            f"Finished at: {backtest_status['finished_at']}",
            f"Result: {backtest_status['result_summary']}",
            f"Error (if any): {backtest_status['error']}",
        ]
        self.wfile.write(("\n".join(lines) + "\n").encode("utf-8"))

    def _download(self):
        if not os.path.exists(OUTPUT_FILE):
            self.send_response(404)
            self.send_header("Content-Type", "text/plain")
            self.end_headers()
            self.wfile.write(b"No backtest_journal.xlsx yet -- visit /run-backtest first.\n")
            return
        try:
            with open(OUTPUT_FILE, "rb") as f:
                data = f.read()
            self.send_response(200)
            self.send_header("Content-Type",
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", "attachment; filename=backtest_journal.xlsx")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
        except Exception as e:
            self.send_response(500)
            self.send_header("Content-Type", "text/plain")
            self.end_headers()
            self.wfile.write(f"Error reading file: {e}\n".encode("utf-8"))

    def log_message(self, format, *args):
        pass


def start_web_server():
    port = int(os.environ.get("PORT", 8080))
    server = HTTPServer(("0.0.0.0", port), _BacktestHandler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    print(f"Backtest web server listening on port {port}.")
    print(f"Visit this service's Render URL to see instructions, "
          f"/run-backtest to start, /status to check, /download when done.")


def main():
    parser = argparse.ArgumentParser(description="Backtest the ORB+VWAP strategy on historical data")
    parser.add_argument("--tickers", type=str, default=None,
                         help="Comma-separated tickers instead of the full Nifty 100")
    parser.add_argument("--days", type=int, default=MAX_LOOKBACK_DAYS,
                         help=f"Lookback days for 5m data (max ~{MAX_LOOKBACK_DAYS}, Yahoo's own limit)")
    parser.add_argument("--once", action="store_true",
                         help="Run once immediately in this terminal and exit (for use via "
                              "Render's Shell tab, or running locally). Without this flag, "
                              "the script instead starts a web server and waits for you to "
                              "trigger a run via /run-backtest -- use this mode when deploying "
                              "as a Render Web Service.")
    args = parser.parse_args()

    if args.once:
        days = min(args.days, MAX_LOOKBACK_DAYS)
        tickers = ([t.strip() for t in args.tickers.split(",") if t.strip()]
                   if args.tickers else get_nifty_100_tickers())

        trades, ending_capital, capital_curve = run_backtest(tickers, days)

        if not trades:
            print("\nNo trades were generated at all -- either the filters are very strict, "
                  "or data fetch failed for most tickers. Check the [warn] lines above.")
            return

        write_excel(trades, BACKTEST_CAPITAL, ending_capital, capital_curve)

        wins = sum(1 for t in trades if "WIN" in t["Outcome"])
        print(f"\n{'='*60}")
        print(f"BACKTEST COMPLETE")
        print(f"  Total trades:     {len(trades)}")
        print(f"  Win rate:         {round(wins/len(trades)*100, 1)}%")
        print(f"  Starting capital: Rs.{BACKTEST_CAPITAL:,.2f}")
        print(f"  Ending capital:   Rs.{ending_capital:,.2f}")
        print(f"  Return:           {round((ending_capital-BACKTEST_CAPITAL)/BACKTEST_CAPITAL*100, 2)}%")
        print(f"  Max drawdown:     {max_drawdown_pct(capital_curve)}%")
        print(f"  Full detail in:   {OUTPUT_FILE}")
        print(f"{'='*60}")
        return

    # Default (no --once): stay alive as a web service, waiting for triggers.
    start_web_server()
    print("Waiting for a backtest to be triggered via /run-backtest. Press Ctrl+C to stop.")
    try:
        while True:
            time.sleep(60)
    except KeyboardInterrupt:
        print("\nStopped by user. Goodbye!")
        sys.exit(0)


if __name__ == "__main__":
    main()
