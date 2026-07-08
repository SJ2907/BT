"""
Breakout + Pullback Strategy BACKTEST (Indian Stock Market - NSE)
=====================================================================
Ported from a Backtrader reference strategy (5Min_Breakout_Pullback_v1)
into plain pandas/Python, so it fits our existing lightweight setup with
no extra heavy dependencies. Fully SEPARATE from the live bot and from
the ORB backtest (BT) -- doesn't touch or import either.

STRATEGY LOGIC (both directions -- the reference only did long/bullish;
a mirrored short/bearish version has been added here for symmetry, noted
clearly since it's an extension beyond the original)
--------------------------------------------------------------------------
1. CONSOLIDATION: track candles whose High-Low range is a small % of price
   (narrow range) for at least MIN_CONSOL_CANDLES in a row.
2. BREAKOUT: out of that consolidation, a candle with an unusually large
   body (vs recent average) AND unusually high volume (vs 20-candle
   average), before MAX_BREAKOUT_HOUR:MAX_BREAKOUT_MINUTE (1:30 PM by
   default -- your "no entries after 1:30 PM" rule lives here).
3. PULLBACK CONFIRMATION: price must pull back toward the breakout level
   (not break too far past it) and then reclaim it on LOW volume within
   PULLBACK_MAX_CANDLES -- that reclaim candle is the actual entry.
4. RISK MANAGEMENT: stop-loss from the consolidation/breakout candle's
   opposite extreme; target = max(2.8R, measured move); stop trails to
   breakeven after 1.2R profit, and trails behind the previous candle's
   high/low afterward. Mandatory flat-by-2:30 PM exit, matching the
   reference's intraday-only design.

IMPORTANT LIMITATIONS -- READ BEFORE TRUSTING THE OUTPUT
--------------------------------------------------------------------------
- Same 60-day cap on 5-minute data as the other backtest (Yahoo's limit,
  not ours) -- see backtest_orb_bot.py's notes for the full explanation.
- No brokerage/STT/slippage modeled.
- Today's Nifty 100 list is used for the whole window (minor survivorship
  bias, same caveat as before).
- NEW SAFETY FEATURE vs the last backtest: MAX_DAILY_RISK_PCT caps how
  much total risk can be committed across ALL simultaneous signals on a
  single day. Without this, a day where many stocks all break out at once
  can silently commit far more than 10% of capital in aggregate -- which
  is what actually wiped out the account in the ORB backtest, more than
  the strategy's own edge or lack of it. Default here is 30% (i.e. at
  most ~3 full-size trades worth of risk open on any given day).
- This is a research simulation, not a guarantee of live results.

REQUIREMENTS
------------
    pip install yfinance pandas pytz openpyxl requests

HOW TO RUN
----------
    python3 breakout_pullback_backtest.py --once                       # run now, in this terminal, then exit
    python3 breakout_pullback_backtest.py --once --tickers RELIANCE.NS,TCS.NS
    python3 breakout_pullback_backtest.py                              # web-service mode (for Render):
                                                                        #   binds a port, waits for a
                                                                        #   trigger at /run-backtest
"""

import argparse
import os
import sys
import threading
import time
from datetime import datetime, timedelta
from http.server import BaseHTTPRequestHandler, HTTPServer
from io import StringIO
from urllib.parse import urlparse, parse_qs

import pandas as pd
import pytz
import requests
import yfinance as yf
from openpyxl import Workbook

# ============================== CONFIG ===================================

BACKTEST_CAPITAL = 300000.0
RISK_PER_TRADE_PCT = 10.0
MAX_DAILY_RISK_PCT = 30.0   # NEW safety cap -- see notes above. Set to e.g.
                            # 1000 to effectively disable it if you want to
                            # see the uncapped/original behavior instead.

CANDLE_INTERVAL = "5m"
MAX_LOOKBACK_DAYS = 59      # Yahoo's real ceiling for 5m data

# --- Strategy parameters (from the reference script) ---
MIN_CONSOL_CANDLES = 8
CONSOL_RANGE_PCT = 1.0
BODY_MULT_BREAKOUT = 1.5
VOL_MULT_BREAKOUT = 2.0
PULLBACK_MAX_CANDLES = 5
PULLBACK_BUFFER_PCT = 0.2
SL_BUFFER_PCT = 0.1
RR1, RR2 = 1.8, 2.8
TRAIL_TO_BE_R = 1.2
MIN_ADR_PCT = 1.2
MAX_BREAKOUT_HOUR, MAX_BREAKOUT_MINUTE = 13, 30   # <-- your "no entry after 1:30 PM" rule
EXIT_HOUR, EXIT_MINUTE = 14, 30                    # mandatory flat-by time

IST = pytz.timezone("Asia/Kolkata")
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "breakout_pullback_journal.xlsx")
NIFTY100_SOURCE_URL = "https://niftyindices.com/IndexConstituent/ind_nifty100list.csv"

FALLBACK_TICKERS = [
    "RELIANCE.NS", "TCS.NS", "HDFCBANK.NS", "ICICIBANK.NS", "INFY.NS",
    "HINDUNILVR.NS", "ITC.NS", "SBIN.NS", "BHARTIARTL.NS", "KOTAKBANK.NS",
]

# ============================ END CONFIG ==================================


def get_nifty_100_tickers():
    try:
        resp = requests.get(NIFTY100_SOURCE_URL,
                             headers={"User-Agent": "Mozilla/5.0"}, timeout=20)
        resp.raise_for_status()
        df = pd.read_csv(StringIO(resp.text))
        symbol_col = "Symbol" if "Symbol" in df.columns else df.columns[2]
        tickers = [f"{s.strip()}.NS" for s in df[symbol_col].dropna().tolist()]
        if len(tickers) < 90:
            raise ValueError(f"Only got {len(tickers)} tickers, expected ~100")
        print(f"Fetched live Nifty 100 list: {len(tickers)} tickers.")
        return tickers
    except Exception as e:
        print(f"[warn] Live Nifty 100 fetch failed ({e}). Using a small fallback list.")
        return FALLBACK_TICKERS


def fetch_intraday_history(ticker, days):
    try:
        df = yf.download(ticker, period=f"{days}d", interval=CANDLE_INTERVAL,
                          progress=False, auto_adjust=False)
    except Exception as e:
        print(f"  [warn] fetch failed for {ticker}: {e}")
        return None
    if df is None or df.empty:
        return None
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    return df


# ------------------------------ STRATEGY CORE ------------------------------

def precompute_indicators(df):
    """Continuous, causal indicators across the WHOLE multi-day series --
    matches how the reference Backtrader feed works (a rolling window can
    dip into the previous day's candles early in a session). Nothing here
    looks ahead in time."""
    body = (df["Close"] - df["Open"])
    df["_avg_body_prior"] = body.abs().shift(1).rolling(10).mean()
    df["_avg_vol"] = df["Volume"].rolling(20).mean()
    df["_recent_range_pct"] = (
        (df["High"].rolling(78).max() - df["Low"].rolling(78).min()) / df["Close"] * 100
    )
    return df


def simulate_ticker(ticker, df, capital_lookup):
    """Runs the full consolidation -> breakout -> pullback -> trade state
    machine across a ticker's ENTIRE historical series. Returns a list of
    raw trade dicts (no position sizing yet -- that happens later, once
    all tickers' signals are pooled together per day for the risk cap)."""
    df = precompute_indicators(df)
    trades = []

    consol_candles = 0
    consol_high = consol_low = None
    in_consolidation = False
    breakout_detected = False
    direction = None
    breakout_level = None
    breakout_vol = None
    breakout_extreme = None  # low (bullish) or high (bearish) of the breakout candle
    pullback_count = 0

    in_position = False
    entry_price = stop_price = trailing_stop = target1 = target2 = measured_move = None
    entry_time = None
    current_day = None

    def reset_setup():
        nonlocal consol_candles, consol_high, consol_low, in_consolidation
        nonlocal breakout_detected, direction, breakout_level, breakout_vol, breakout_extreme, pullback_count
        consol_candles = 0
        consol_high = consol_low = None
        in_consolidation = False
        breakout_detected = False
        direction = None
        breakout_level = None
        breakout_vol = None
        breakout_extreme = None
        pullback_count = 0

    n = len(df)
    for i in range(n):
        row = df.iloc[i]
        row_time = row.name.time()
        row_day = row.name.date()

        if row_day != current_day:
            current_day = row_day
            reset_setup()  # intraday-only strategy: nothing carries across days
            in_position = False  # should already be flat from EOD exit; safety net

        if i < 25:  # warm-up, like the reference's `len(self) < 50` guard (roughly)
            continue

        if in_position:
            # Mandatory flat-by-time exit
            if row_time >= pd.Timestamp(f"{EXIT_HOUR}:{EXIT_MINUTE}").time():
                exit_price = float(row["Close"])
                trades.append(_build_trade(ticker, direction, entry_time, entry_price,
                                            stop_price, target1, target2, row.name, exit_price,
                                            "EOD_SQUAREOFF"))
                in_position = False
                reset_setup()
                continue

            price = float(row["Close"])
            risk = abs(entry_price - stop_price)
            prev_low = float(df.iloc[i - 1]["Low"])
            prev_high = float(df.iloc[i - 1]["High"])

            if direction == "BULLISH":
                if price >= max(target2, measured_move):
                    trades.append(_build_trade(ticker, direction, entry_time, entry_price,
                                                stop_price, target1, target2, row.name, price,
                                                "TARGET_HIT"))
                    in_position = False
                    reset_setup()
                    continue
                if (price - entry_price) / risk >= TRAIL_TO_BE_R:
                    trailing_stop = max(trailing_stop, entry_price)
                trailing_stop = max(trailing_stop, prev_low)
                if price <= trailing_stop:
                    trades.append(_build_trade(ticker, direction, entry_time, entry_price,
                                                stop_price, target1, target2, row.name, price,
                                                "TRAIL_STOP_HIT"))
                    in_position = False
                    reset_setup()
                    continue
            else:  # BEARISH
                if price <= min(target2, measured_move):
                    trades.append(_build_trade(ticker, direction, entry_time, entry_price,
                                                stop_price, target1, target2, row.name, price,
                                                "TARGET_HIT"))
                    in_position = False
                    reset_setup()
                    continue
                if (entry_price - price) / risk >= TRAIL_TO_BE_R:
                    trailing_stop = min(trailing_stop, entry_price)
                trailing_stop = min(trailing_stop, prev_high)
                if price >= trailing_stop:
                    trades.append(_build_trade(ticker, direction, entry_time, entry_price,
                                                stop_price, target1, target2, row.name, price,
                                                "TRAIL_STOP_HIT"))
                    in_position = False
                    reset_setup()
                    continue
            continue  # while in a position, skip new-setup detection this candle

        # --- Not in a position: build/monitor the setup ---
        recent_range = row["_recent_range_pct"]
        if pd.isna(recent_range) or recent_range < MIN_ADR_PCT:
            reset_setup()
            continue

        candle_range_pct = (row["High"] - row["Low"]) / row["Close"] * 100
        is_narrow = candle_range_pct <= CONSOL_RANGE_PCT
        early_enough = row_time <= pd.Timestamp(f"{MAX_BREAKOUT_HOUR}:{MAX_BREAKOUT_MINUTE}").time()

        # IMPORTANT FIX vs the reference script: check for a breakout using the
        # consolidation state as it stood BEFORE this candle, not after. A
        # breakout candle is by definition wide-range -- if we updated the
        # consolidation counters with this candle FIRST, its own wideness
        # would reset in_consolidation to False right before ever checking
        # whether it broke out of that consolidation, so a breakout could
        # almost never be detected. Checking first, then updating, fixes this.
        just_detected_breakout = False
        if in_consolidation and not breakout_detected and early_enough:
            body = row["Close"] - row["Open"]
            avg_body = row["_avg_body_prior"]
            avg_vol = row["_avg_vol"]
            if pd.notna(avg_body) and pd.notna(avg_vol):
                strong_body = abs(body) >= BODY_MULT_BREAKOUT * avg_body if avg_body > 0 else abs(body) > 0
                high_volume = row["Volume"] >= VOL_MULT_BREAKOUT * avg_vol
                if body > 0 and strong_body and high_volume:
                    direction = "BULLISH"
                    breakout_detected = True
                    just_detected_breakout = True
                    breakout_level = max(consol_high, row["High"])
                    breakout_vol = row["Volume"]
                    breakout_extreme = row["Low"]
                    in_consolidation = False
                elif body < 0 and strong_body and high_volume:
                    direction = "BEARISH"
                    breakout_detected = True
                    just_detected_breakout = True
                    breakout_level = min(consol_low, row["Low"])
                    breakout_vol = row["Volume"]
                    breakout_extreme = row["High"]
                    in_consolidation = False

        # Now update consolidation tracking for the NEXT candle's use
        if not breakout_detected:
            if is_narrow:
                consol_candles += 1
                consol_high = row["High"] if consol_high is None else max(consol_high, row["High"])
                consol_low = row["Low"] if consol_low is None else min(consol_low, row["Low"])
                if consol_candles >= MIN_CONSOL_CANDLES:
                    in_consolidation = True
            else:
                consol_candles, consol_high, consol_low, in_consolidation = 0, None, None, False

        if just_detected_breakout:
            # ANOTHER FIX vs the reference script: pullback monitoring must
            # start on the candle AFTER the breakout, not the same one. A
            # breakout candle's own low is, almost by definition, well below
            # the breakout level -- checking the invalidation condition
            # against that same candle would nearly always instantly
            # invalidate the setup before a real pullback ever gets a chance.
            continue

        if breakout_detected and not in_position:
            prev_close = float(df.iloc[i - 1]["Close"])
            if direction == "BULLISH":
                if row["Low"] < breakout_level * (1 - PULLBACK_BUFFER_PCT / 100):
                    reset_setup()
                    continue
                pullback_count += 1
                if 1 <= pullback_count <= PULLBACK_MAX_CANDLES:
                    confirm = (row["Close"] > row["Open"] and row["Close"] > breakout_level
                               and prev_close <= breakout_level)
                    low_vol = row["Volume"] < 0.7 * breakout_vol
                    if confirm and low_vol:
                        entry_price = float(row["Close"])
                        stop_price = min(consol_low, breakout_extreme) * (1 - SL_BUFFER_PCT / 100)
                        risk = entry_price - stop_price
                        if risk > 0:
                            trailing_stop = stop_price
                            target1 = entry_price + RR1 * risk
                            target2 = entry_price + RR2 * risk
                            measured_move = breakout_level + (breakout_level - consol_low)
                            entry_time = row.name
                            in_position = True
                        else:
                            reset_setup()
                if pullback_count > PULLBACK_MAX_CANDLES:
                    reset_setup()
            else:  # BEARISH
                if row["High"] > breakout_level * (1 + PULLBACK_BUFFER_PCT / 100):
                    reset_setup()
                    continue
                pullback_count += 1
                if 1 <= pullback_count <= PULLBACK_MAX_CANDLES:
                    confirm = (row["Close"] < row["Open"] and row["Close"] < breakout_level
                               and prev_close >= breakout_level)
                    low_vol = row["Volume"] < 0.7 * breakout_vol
                    if confirm and low_vol:
                        entry_price = float(row["Close"])
                        stop_price = max(consol_high, breakout_extreme) * (1 + SL_BUFFER_PCT / 100)
                        risk = stop_price - entry_price
                        if risk > 0:
                            trailing_stop = stop_price
                            target1 = entry_price - RR1 * risk
                            target2 = entry_price - RR2 * risk
                            measured_move = breakout_level - (consol_high - breakout_level)
                            entry_time = row.name
                            in_position = True
                        else:
                            reset_setup()
                if pullback_count > PULLBACK_MAX_CANDLES:
                    reset_setup()

    return trades


def _build_trade(ticker, direction, entry_time, entry_price, stop_price, target1, target2,
                  exit_time, exit_price, exit_reason):
    risk = abs(entry_price - stop_price)
    return {
        "EntryDate": entry_time.strftime("%Y-%m-%d"),
        "EntryTime": entry_time.strftime("%H:%M:%S"),
        "Ticker": ticker, "Direction": direction,
        "Entry": round(entry_price, 2), "InitialStop": round(stop_price, 2),
        "Target1": round(target1, 2), "Target2": round(target2, 2),
        "RiskPerShare": round(risk, 2),
        "ExitTime": exit_time.strftime("%H:%M:%S"), "ExitPrice": round(exit_price, 2),
        "ExitReason": exit_reason,
    }


# ------------------------------ POSITION SIZING -----------------------------

def apply_position_sizing(raw_trades):
    """Pools ALL tickers' raw signals together, processes them in
    chronological order day-by-day, and applies capital-based sizing plus
    the daily risk cap -- this is the step that prevents a cluster of
    same-day signals from silently over-committing capital."""
    by_day = {}
    for t in raw_trades:
        by_day.setdefault(t["EntryDate"], []).append(t)

    capital = BACKTEST_CAPITAL
    final_trades = []
    capital_curve = [capital]

    for day in sorted(by_day.keys()):
        day_trades = sorted(by_day[day], key=lambda t: t["EntryTime"])
        capital_start_of_day = capital
        committed_risk = 0.0
        day_pnl = 0.0
        taken_count = 0

        for t in day_trades:
            risk_amount = capital_start_of_day * (RISK_PER_TRADE_PCT / 100)
            if committed_risk + risk_amount > capital_start_of_day * (MAX_DAILY_RISK_PCT / 100):
                continue  # daily risk cap reached -- skip remaining signals today
            committed_risk += risk_amount

            qty = max(1, int(risk_amount // t["RiskPerShare"])) if t["RiskPerShare"] > 0 else 0
            if qty == 0:
                continue

            if t["Direction"] == "BULLISH":
                pnl = (t["ExitPrice"] - t["Entry"]) * qty
            else:
                pnl = (t["Entry"] - t["ExitPrice"]) * qty
            pnl_pct = round((pnl / (t["Entry"] * qty)) * 100, 2) if t["Entry"] * qty else 0
            outcome = "WIN" if pnl > 0 else "LOSS"

            t["Qty"] = qty
            t["RiskAmount"] = round(risk_amount, 2)
            t["PnL"] = round(pnl, 2)
            t["PnLPct"] = pnl_pct
            t["Outcome"] = outcome
            day_pnl += pnl
            taken_count += 1
            final_trades.append(t)

        if taken_count > 0:
            capital += day_pnl
            for t in final_trades[-taken_count:]:
                t["CapitalAfter"] = round(capital, 2)
            capital_curve.append(capital)
            skipped = len(day_trades) - taken_count
            skip_note = f" ({skipped} skipped by daily risk cap)" if skipped else ""
            print(f"  {day}: {taken_count} trade(s) taken{skip_note}, "
                  f"day P&L Rs.{round(day_pnl,2)}, capital now Rs.{round(capital,2)}")

    return final_trades, capital, capital_curve


def run_backtest(tickers, days):
    print(f"Fetching {days}-day intraday history for {len(tickers)} tickers...")
    all_raw_trades = []
    usable = 0
    for i, ticker in enumerate(tickers, 1):
        print(f"  [{i}/{len(tickers)}] {ticker}...")
        df = fetch_intraday_history(ticker, days)
        if df is None or len(df) < 100:
            continue
        usable += 1
        try:
            raw = simulate_ticker(ticker, df, None)
            all_raw_trades.extend(raw)
        except Exception as e:
            print(f"  [warn] simulation error for {ticker}: {e}")

    print(f"\nGot usable data for {usable}/{len(tickers)} tickers. "
          f"{len(all_raw_trades)} raw signals found. Applying position sizing...")

    return apply_position_sizing(all_raw_trades)


# ------------------------------ EXCEL OUTPUT --------------------------------

TRADE_COLUMNS = ["EntryDate", "EntryTime", "Ticker", "Direction", "Entry", "InitialStop",
                  "Target1", "Target2", "Qty", "RiskAmount", "ExitTime", "ExitPrice",
                  "ExitReason", "Outcome", "PnL", "PnLPct", "CapitalAfter"]


def max_drawdown_pct(curve):
    peak = curve[0]
    max_dd = 0.0
    for value in curve:
        peak = max(peak, value)
        dd = (peak - value) / peak * 100 if peak else 0
        max_dd = max(max_dd, dd)
    return round(max_dd, 2)


def write_excel(trades, starting_capital, ending_capital, capital_curve):
    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"
    ws.append(TRADE_COLUMNS)
    for t in trades:
        ws.append([t.get(c, "") for c in TRADE_COLUMNS])

    wins = sum(1 for t in trades if t["Outcome"] == "WIN")
    total = len(trades)
    win_rate = round(wins / total * 100, 2) if total else 0
    total_pnl = round(sum(t["PnL"] for t in trades), 2)
    return_pct = round((ending_capital - starting_capital) / starting_capital * 100, 2)

    summary_ws = wb.create_sheet("Summary")
    for row in [
        ("Starting Capital", starting_capital),
        ("Ending Capital", round(ending_capital, 2)),
        ("Total Return %", return_pct),
        ("Total Trades", total),
        ("Wins", wins),
        ("Losses", total - wins),
        ("Win Rate %", win_rate),
        ("Total P&L", total_pnl),
        ("Max Drawdown %", max_drawdown_pct(capital_curve)),
        ("Risk Per Trade %", RISK_PER_TRADE_PCT),
        ("Max Daily Risk Cap %", MAX_DAILY_RISK_PCT),
        ("No entries after", f"{MAX_BREAKOUT_HOUR}:{MAX_BREAKOUT_MINUTE:02d}"),
        ("Mandatory exit by", f"{EXIT_HOUR}:{EXIT_MINUTE:02d}"),
    ]:
        summary_ws.append(row)

    by_symbol = {}
    for t in trades:
        s = by_symbol.setdefault(t["Ticker"], {"trades": 0, "wins": 0, "pnl": 0.0})
        s["trades"] += 1
        s["wins"] += 1 if t["Outcome"] == "WIN" else 0
        s["pnl"] += t["PnL"]
    symbol_ws = wb.create_sheet("BySymbol")
    symbol_ws.append(["Ticker", "Trades", "Wins", "WinRate%", "TotalPnL"])
    for ticker, s in sorted(by_symbol.items(), key=lambda x: -x[1]["pnl"]):
        wr = round(s["wins"] / s["trades"] * 100, 2) if s["trades"] else 0
        symbol_ws.append([ticker, s["trades"], s["wins"], wr, round(s["pnl"], 2)])

    wb.save(OUTPUT_FILE)


# --------------------------- STANDALONE WEB SERVER --------------------------

backtest_status = {"running": False, "started_at": None, "finished_at": None,
                    "result_summary": None, "error": None}


def _run_backtest_background(days, tickers_arg):
    backtest_status["running"] = True
    backtest_status["started_at"] = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    backtest_status["error"] = None
    try:
        tickers = ([t.strip() for t in tickers_arg.split(",") if t.strip()]
                   if tickers_arg else get_nifty_100_tickers())
        days = min(days, MAX_LOOKBACK_DAYS)
        trades, ending_capital, curve = run_backtest(tickers, days)
        if trades:
            write_excel(trades, BACKTEST_CAPITAL, ending_capital, curve)
            wins = sum(1 for t in trades if t["Outcome"] == "WIN")
            backtest_status["result_summary"] = (
                f"{len(trades)} trades, win rate {round(wins/len(trades)*100,1)}%, "
                f"ended at Rs.{round(ending_capital,2)} (started Rs.{BACKTEST_CAPITAL:,.2f}), "
                f"max drawdown {max_drawdown_pct(curve)}%"
            )
        else:
            backtest_status["result_summary"] = "No trades were generated -- check server logs."
    except Exception as e:
        backtest_status["error"] = str(e)
    finally:
        backtest_status["running"] = False
        backtest_status["finished_at"] = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")


class _Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        if self.path.startswith("/run-backtest"):
            self._trigger()
        elif self.path.startswith("/status"):
            self._status()
        elif self.path.startswith("/download"):
            self._download()
        else:
            self._home()

    def do_HEAD(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()

    def _home(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()
        self.wfile.write((
            "Breakout-Pullback backtest service alive.\n\n"
            "Visit /run-backtest to start (optional: ?days=30&tickers=RELIANCE.NS,TCS.NS)\n"
            "Visit /status to check progress.\n"
            "Visit /download to get breakout_pullback_journal.xlsx once finished.\n"
        ).encode("utf-8"))

    def _trigger(self):
        if backtest_status["running"]:
            self.send_response(200)
            self.send_header("Content-Type", "text/plain")
            self.end_headers()
            self.wfile.write(b"A backtest is already running. Check /status.\n")
            return
        query = parse_qs(urlparse(self.path).query)
        days = int(query.get("days", [MAX_LOOKBACK_DAYS])[0])
        tickers_arg = query.get("tickers", [None])[0]
        threading.Thread(target=_run_backtest_background, args=(days, tickers_arg), daemon=True).start()
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()
        self.wfile.write(
            f"Backtest started ({days} days). Check /status for progress, /download when done.\n".encode("utf-8")
        )

    def _status(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()
        lines = [f"Running: {backtest_status['running']}",
                 f"Started at: {backtest_status['started_at']}",
                 f"Finished at: {backtest_status['finished_at']}",
                 f"Result: {backtest_status['result_summary']}",
                 f"Error (if any): {backtest_status['error']}"]
        self.wfile.write(("\n".join(lines) + "\n").encode("utf-8"))

    def _download(self):
        if not os.path.exists(OUTPUT_FILE):
            self.send_response(404)
            self.send_header("Content-Type", "text/plain")
            self.end_headers()
            self.wfile.write(b"No breakout_pullback_journal.xlsx yet -- visit /run-backtest first.\n")
            return
        with open(OUTPUT_FILE, "rb") as f:
            data = f.read()
        self.send_response(200)
        self.send_header("Content-Type",
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", "attachment; filename=breakout_pullback_journal.xlsx")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, format, *args):
        pass


def start_web_server():
    port = int(os.environ.get("PORT", 8080))
    server = HTTPServer(("0.0.0.0", port), _Handler)
    threading.Thread(target=server.serve_forever, daemon=True).start()
    print(f"Breakout-Pullback backtest server listening on port {port}.")


def main():
    parser = argparse.ArgumentParser(description="Backtest the breakout-pullback strategy")
    parser.add_argument("--tickers", type=str, default=None)
    parser.add_argument("--days", type=int, default=MAX_LOOKBACK_DAYS)
    parser.add_argument("--once", action="store_true",
                         help="Run once immediately and exit (Shell tab / local use). "
                              "Without this flag, starts a web server for Render deployment.")
    args = parser.parse_args()

    if args.once:
        days = min(args.days, MAX_LOOKBACK_DAYS)
        tickers = ([t.strip() for t in args.tickers.split(",") if t.strip()]
                   if args.tickers else get_nifty_100_tickers())
        trades, ending_capital, curve = run_backtest(tickers, days)
        if not trades:
            print("\nNo trades were generated -- check the [warn] lines above.")
            return
        write_excel(trades, BACKTEST_CAPITAL, ending_capital, curve)
        wins = sum(1 for t in trades if t["Outcome"] == "WIN")
        print(f"\n{'='*60}\nBACKTEST COMPLETE")
        print(f"  Total trades:     {len(trades)}")
        print(f"  Win rate:         {round(wins/len(trades)*100, 1)}%")
        print(f"  Starting capital: Rs.{BACKTEST_CAPITAL:,.2f}")
        print(f"  Ending capital:   Rs.{ending_capital:,.2f}")
        print(f"  Return:           {round((ending_capital-BACKTEST_CAPITAL)/BACKTEST_CAPITAL*100, 2)}%")
        print(f"  Max drawdown:     {max_drawdown_pct(curve)}%")
        print(f"  Full detail in:   {OUTPUT_FILE}\n{'='*60}")
        return

    start_web_server()
    print("Waiting for a backtest to be triggered via /run-backtest. Press Ctrl+C to stop.")
    try:
        while True:
            time.sleep(60)
    except KeyboardInterrupt:
        print("\nStopped by user. Goodbye!")
        sys.exit(0)


if __name__ == "__main__":
    main()
