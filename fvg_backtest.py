"""
Box Rejection ("Wick") Strategy BACKTEST -- Previous-Day Top-20 Gainers
=========================================================================
(Indian Stock Market - NSE)

Fully SEPARATE from the live bot and every other backtest script here.
Doesn't touch or import any of them.

Adapted from a reference "Box Theory" script, but rebuilt on our own
validated backtest framework -- the reference's own backtest logic never
actually recorded real per-trade P&L; its "equity curve" was a disconnected
approximation that didn't match its own entry/exit rules. This version
uses proper trade-by-trade simulation with real position sizing, matching
the ORB/breakout-pullback/FVG/BAG backtests already built.

KEY DIFFERENCE FROM THE OTHER BACKTESTS: the tradeable universe ROTATES
DAILY. Instead of scanning the full Nifty 100 every day, each day only
scans that day's TOP_N_GAINERS stocks (by previous day's % gain, ranked
across the Nifty 100). The idea: stocks with recent relative strength may
behave differently at support/resistance than a static universe would
capture.

STRATEGY LOGIC
--------------
1. THE "BOX": previous trading day's High and Low act as a support/
   resistance zone for today's session.

2. THE "WICK" (rejection candle):
     Bullish: a candle whose LOWER wick is a large majority (>=WICK_RATIO)
              of its total range, with the low touching at or near the
              box's low (within BOX_TOUCH_BUFFER_PCT) -- price dipped to
              yesterday's low, got rejected, and closed back up.
     Bearish: the mirror -- long upper wick rejecting at/near yesterday's
              high.

3. ENTRY: at the rejection candle's own close. Stop just beyond that same
   candle's extreme (its low, for a bullish rejection). Target = a fixed
   R-multiple of that risk (TARGET_RR, default 2.0, matching the
   reference). Time-boxed: exit after MAX_HOLD_CANDLES if neither stop nor
   target is hit, in addition to the usual mandatory EOD square-off.

IMPORTANT LIMITATIONS
--------------------------------------------------------------------------
- Same ~60-day cap on 5-minute data as the other backtests (Yahoo's limit).
- No brokerage/STT/slippage modeled.
- Needs at least 2 prior trading days of daily closes before the first
  tradeable day, to compute "previous day's % gain" for ranking -- so the
  very start of the backtest window has no ranking yet.
- Today's Nifty 100 list is used as the ranking universe for the whole
  window (same minor survivorship-bias caveat as the other backtests).
- MAX_DAILY_RISK_PCT caps total risk committed across all simultaneous
  signals on one day.
- Exits check High/Low for actual level touches (not just candle Close),
  matching the fix applied to the other backtests -- avoids overstating
  losses from a stop that gets "noticed" several candles late.
- Research simulation only, not a guarantee of live results.

REQUIREMENTS
------------
    pip install yfinance pandas pytz openpyxl requests

HOW TO RUN
----------
    python3 box_wick_backtest.py --once                     # run now, in this terminal, then exit
    python3 box_wick_backtest.py --once --days 30
    python3 box_wick_backtest.py                             # web-service mode (for Render):
                                                              #   binds a port, waits for a
                                                              #   trigger at /run-backtest
"""

import argparse
import os
import sys
import threading
import time
import uuid
from datetime import datetime, timedelta
from http.server import BaseHTTPRequestHandler, HTTPServer
from io import StringIO
from urllib.parse import urlparse, parse_qs

import pandas as pd
import pytz
import requests
import yfinance as yf
from openpyxl import Workbook

# ============================== CONFIG ===================================

BACKTEST_CAPITAL = 300000.0
RISK_PER_TRADE_PCT = 10.0
MAX_DAILY_RISK_PCT = 30.0

CANDLE_INTERVAL = "5m"
MAX_LOOKBACK_DAYS = 59

TOP_N_GAINERS = 10          # how many previous-day top gainers to scan each day

WICK_RATIO = 0.6            # rejection wick must be >= this fraction of the candle's range
BOX_TOUCH_BUFFER_PCT = 0.2  # how close to the box level counts as "touching"
SL_BUFFER_PCT = 0.2         # stop placed this % beyond the rejection candle's own extreme
TARGET_RR = 2.0
MAX_HOLD_CANDLES = 12       # ~1 hour on 5-min candles, matching the reference's default

NO_ENTRY_AFTER_HOUR, NO_ENTRY_AFTER_MINUTE = 13, 30   # 1:30 PM -- leaves room for a
                                                        # full 1-hour hold before the
                                                        # mandatory 2:30 PM close below
EXIT_HOUR, EXIT_MINUTE = 14, 30                        # mandatory flat-by 2:30 PM

IST = pytz.timezone("Asia/Kolkata")
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "box_wick_journal.xlsx")
NIFTY100_SOURCE_URL = "https://niftyindices.com/IndexConstituent/ind_nifty100list.csv"

FALLBACK_TICKERS = [
    "RELIANCE.NS", "TCS.NS", "HDFCBANK.NS", "ICICIBANK.NS", "INFY.NS",
    "HINDUNILVR.NS", "ITC.NS", "SBIN.NS", "BHARTIARTL.NS", "KOTAKBANK.NS",
    "LT.NS", "BAJFINANCE.NS", "HCLTECH.NS", "AXISBANK.NS", "ASIANPAINT.NS",
    "MARUTI.NS", "SUNPHARMA.NS", "TITAN.NS", "ULTRACEMCO.NS", "WIPRO.NS",
]

# ============================ END CONFIG ==================================


def get_nifty_100_tickers():
    try:
        resp = requests.get(NIFTY100_SOURCE_URL,
                             headers={"User-Agent": "Mozilla/5.0"}, timeout=20)
        resp.raise_for_status()
        df = pd.read_csv(StringIO(resp.text))
        symbol_col = "Symbol" if "Symbol" in df.columns else df.columns[2]
        tickers = [f"{s.strip()}.NS" for s in df[symbol_col].dropna().tolist()]
        if len(tickers) < 90:
            raise ValueError(f"Only got {len(tickers)} tickers, expected ~100")
        print(f"Fetched live Nifty 100 list: {len(tickers)} tickers.")
        return tickers
    except Exception as e:
        print(f"[warn] Live Nifty 100 fetch failed ({e}). Using a small fallback list.")
        return FALLBACK_TICKERS


def fetch_intraday_history(ticker, days):
    try:
        df = yf.download(ticker, period=f"{days}d", interval=CANDLE_INTERVAL,
                          progress=False, auto_adjust=False)
    except Exception as e:
        print(f"  [warn] intraday fetch failed for {ticker}: {e}")
        return None
    if df is None or df.empty:
        return None
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    return df


def fetch_daily_history(ticker, days):
    try:
        df = yf.download(ticker, period=f"{days + 15}d", interval="1d",
                          progress=False, auto_adjust=False)
    except Exception as e:
        print(f"  [warn] daily fetch failed for {ticker}: {e}")
        return None
    if df is None or df.empty:
        return None
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    return df


# ------------------------------ RANKING -------------------------------------

def build_daily_rankings(daily_data):
    """For each day present in the data, ranks tickers by their % gain on
    the PRIOR trading day (Close-to-Close), and returns
    {date: [ticker1, ticker2, ...]} with only the top TOP_N_GAINERS kept.
    Only uses data strictly before `date` -- no look-ahead."""
    # Build a single {ticker: {date: close}} lookup
    closes = {}
    all_dates = set()
    for ticker, df in daily_data.items():
        closes[ticker] = {d.date(): float(c) for d, c in zip(df.index, df["Close"])}
        all_dates.update(closes[ticker].keys())

    sorted_dates = sorted(all_dates)
    rankings = {}

    for idx, today in enumerate(sorted_dates):
        if idx < 2:
            continue  # need at least 2 prior closes to compute yesterday's % gain
        yesterday = sorted_dates[idx - 1]
        day_before = sorted_dates[idx - 2]

        gains = []
        for ticker, close_map in closes.items():
            if yesterday in close_map and day_before in close_map and close_map[day_before] > 0:
                pct_gain = (close_map[yesterday] - close_map[day_before]) / close_map[day_before] * 100
                gains.append((ticker, pct_gain))

        gains.sort(key=lambda x: -x[1])
        rankings[today] = [t for t, _ in gains[:TOP_N_GAINERS]]

    return rankings


# ------------------------------ STRATEGY CORE ------------------------------

def is_bullish_wick(row, box_low):
    total_range = row["High"] - row["Low"]
    if total_range <= 0:
        return False
    lower_wick = min(row["Open"], row["Close"]) - row["Low"]
    return (lower_wick / total_range >= WICK_RATIO) and (row["Low"] <= box_low * (1 + BOX_TOUCH_BUFFER_PCT / 100))


def is_bearish_wick(row, box_high):
    total_range = row["High"] - row["Low"]
    if total_range <= 0:
        return False
    upper_wick = row["High"] - max(row["Open"], row["Close"])
    return (upper_wick / total_range >= WICK_RATIO) and (row["High"] >= box_high * (1 - BOX_TOUCH_BUFFER_PCT / 100))


def _time_at_or_after(row_time, hour, minute):
    return row_time >= pd.Timestamp(f"{hour}:{minute}").time()


def simulate_ticker_day(ticker, day_df, box_high, box_low):
    """Looks for ONE wick-rejection trade on this ticker's single day of
    candles. Returns (raw_trade_or_None, all_signals_this_day)."""
    all_signals = []
    n = len(day_df)

    for i in range(n):
        row = day_df.iloc[i]
        row_time = row.name.time()
        if _time_at_or_after(row_time, NO_ENTRY_AFTER_HOUR, NO_ENTRY_AFTER_MINUTE):
            break

        bullish = is_bullish_wick(row, box_low)
        bearish = is_bearish_wick(row, box_high)
        if not bullish and not bearish:
            continue

        direction = "BULLISH" if bullish else "BEARISH"
        sig_id = str(uuid.uuid4())[:8]
        all_signals.append({
            "SignalId": sig_id, "Date": row.name.strftime("%Y-%m-%d"),
            "DetectedTime": row.name.strftime("%H:%M:%S"), "Ticker": ticker,
            "Direction": direction, "BoxHigh": round(box_high, 2), "BoxLow": round(box_low, 2),
            "EntryTriggered": True, "EntryTime": row.name.strftime("%H:%M:%S"),
        })

        entry_price = float(row["Close"])
        if direction == "BULLISH":
            stop_price = float(row["Low"]) * (1 - SL_BUFFER_PCT / 100)
            risk = entry_price - stop_price
            if risk <= 0:
                continue
            target = entry_price + TARGET_RR * risk
        else:
            stop_price = float(row["High"]) * (1 + SL_BUFFER_PCT / 100)
            risk = stop_price - entry_price
            if risk <= 0:
                continue
            target = entry_price - TARGET_RR * risk

        entry_time = row.name
        entry_idx = i

        # Walk forward: check High/Low touches (not just Close), a
        # max-hold-candles cap, and the mandatory EOD square-off.
        exit_price, exit_time, exit_reason = None, None, None
        for j in range(entry_idx + 1, n):
            frow = day_df.iloc[j]
            frow_time = frow.name.time()

            if _time_at_or_after(frow_time, EXIT_HOUR, EXIT_MINUTE):
                exit_price, exit_time, exit_reason = float(frow["Close"]), frow.name, "EOD_SQUAREOFF"
                break

            if direction == "BULLISH":
                if frow["Open"] <= stop_price:
                    exit_price, exit_time, exit_reason = float(frow["Open"]), frow.name, "STOP_HIT"
                    break
                if frow["Low"] <= stop_price:
                    exit_price, exit_time, exit_reason = stop_price, frow.name, "STOP_HIT"
                    break
                if frow["Open"] >= target:
                    exit_price, exit_time, exit_reason = float(frow["Open"]), frow.name, "TARGET_HIT"
                    break
                if frow["High"] >= target:
                    exit_price, exit_time, exit_reason = target, frow.name, "TARGET_HIT"
                    break
            else:
                if frow["Open"] >= stop_price:
                    exit_price, exit_time, exit_reason = float(frow["Open"]), frow.name, "STOP_HIT"
                    break
                if frow["High"] >= stop_price:
                    exit_price, exit_time, exit_reason = stop_price, frow.name, "STOP_HIT"
                    break
                if frow["Open"] <= target:
                    exit_price, exit_time, exit_reason = float(frow["Open"]), frow.name, "TARGET_HIT"
                    break
                if frow["Low"] <= target:
                    exit_price, exit_time, exit_reason = target, frow.name, "TARGET_HIT"
                    break

            if (j - entry_idx) >= MAX_HOLD_CANDLES:
                exit_price, exit_time, exit_reason = float(frow["Close"]), frow.name, "MAX_HOLD_TIME"
                break

        if exit_price is None:
            # ran out of candles today without any exit condition firing
            last_row = day_df.iloc[-1]
            exit_price, exit_time, exit_reason = float(last_row["Close"]), last_row.name, "EOD_SQUAREOFF"

        trade = {
            "SignalId": sig_id,
            "EntryDate": entry_time.strftime("%Y-%m-%d"), "EntryTime": entry_time.strftime("%H:%M:%S"),
            "Ticker": ticker, "Direction": direction,
            "Entry": round(entry_price, 2), "InitialStop": round(stop_price, 2),
            "Target": round(target, 2), "RiskPerShare": round(risk, 2),
            "ExitTime": exit_time.strftime("%H:%M:%S"), "ExitPrice": round(exit_price, 2),
            "ExitReason": exit_reason,
        }
        return trade, all_signals  # only one trade per ticker per day

    return None, all_signals


# ------------------------------ POSITION SIZING -----------------------------

def apply_position_sizing(raw_trades):
    by_day = {}
    for t in raw_trades:
        by_day.setdefault(t["EntryDate"], []).append(t)

    capital = BACKTEST_CAPITAL
    final_trades = []
    capital_curve = [capital]
    taken_signal_ids = set()

    for day in sorted(by_day.keys()):
        day_trades = sorted(by_day[day], key=lambda t: t["EntryTime"])
        capital_start_of_day = capital
        committed_risk = 0.0
        day_pnl = 0.0
        taken_count = 0

        for t in day_trades:
            risk_amount = capital_start_of_day * (RISK_PER_TRADE_PCT / 100)
            if committed_risk + risk_amount > capital_start_of_day * (MAX_DAILY_RISK_PCT / 100):
                continue
            committed_risk += risk_amount
            qty = max(1, int(risk_amount // t["RiskPerShare"])) if t["RiskPerShare"] > 0 else 0
            if qty == 0:
                continue

            if t["Direction"] == "BULLISH":
                pnl = (t["ExitPrice"] - t["Entry"]) * qty
            else:
                pnl = (t["Entry"] - t["ExitPrice"]) * qty
            pnl_pct = round((pnl / (t["Entry"] * qty)) * 100, 2) if t["Entry"] * qty else 0
            outcome = "WIN" if pnl > 0 else "LOSS"

            t["Qty"] = qty
            t["RiskAmount"] = round(risk_amount, 2)
            t["PnL"] = round(pnl, 2)
            t["PnLPct"] = pnl_pct
            t["Outcome"] = outcome
            day_pnl += pnl
            taken_count += 1
            taken_signal_ids.add(t["SignalId"])
            final_trades.append(t)

        if taken_count > 0:
            capital += day_pnl
            for t in final_trades[-taken_count:]:
                t["CapitalAfter"] = round(capital, 2)
            capital_curve.append(capital)
            skipped = len(day_trades) - taken_count
            skip_note = f" ({skipped} skipped by daily risk cap)" if skipped else ""
            print(f"  {day}: {taken_count} trade(s) taken{skip_note}, "
                  f"day P&L Rs.{round(day_pnl,2)}, capital now Rs.{round(capital,2)}")

    return final_trades, capital, capital_curve, taken_signal_ids


def run_backtest(days):
    tickers = get_nifty_100_tickers()
    print(f"Fetching {days}-day intraday + daily history for {len(tickers)} tickers "
          f"(needed to rank top-{TOP_N_GAINERS} gainers each day)...")

    intraday_data, daily_data = {}, {}
    for i, ticker in enumerate(tickers, 1):
        print(f"  [{i}/{len(tickers)}] {ticker}...")
        idf = fetch_intraday_history(ticker, days)
        ddf = fetch_daily_history(ticker, days)
        if idf is not None and ddf is not None:
            intraday_data[ticker] = idf
            daily_data[ticker] = ddf

    print(f"\nGot usable data for {len(intraday_data)}/{len(tickers)} tickers. "
          f"Building daily top-{TOP_N_GAINERS} rankings...")
    rankings = build_daily_rankings(daily_data)

    all_raw_trades, all_signals = [], []
    for day, top_tickers in sorted(rankings.items()):
        # ONLY ONE TRADE PER DAY, TOTAL -- scan the top gainers in rank order
        # (strongest gainer first) and stop as soon as one produces a trade.
        # Tickers ranked below the one that fires are never even checked
        # that day, matching "find my one trade for the day and stop looking."
        for ticker in top_tickers:
            idf = intraday_data.get(ticker)
            ddf = daily_data.get(ticker)
            if idf is None or ddf is None:
                continue
            day_df = idf[idf.index.date == day].copy()
            if day_df.empty:
                continue
            prior_days = ddf[ddf.index.date < day]
            if prior_days.empty:
                continue
            prev = prior_days.iloc[-1]
            box_high, box_low = float(prev["High"]), float(prev["Low"])

            try:
                trade, signals = simulate_ticker_day(ticker, day_df, box_high, box_low)
            except Exception as e:
                print(f"  [warn] simulation error for {ticker} on {day}: {e}")
                continue

            all_signals.extend(signals)
            if trade:
                all_raw_trades.append(trade)
                break  # got today's one trade -- stop scanning remaining tickers

    print(f"\n{len(all_signals)} wick rejections detected, {len(all_raw_trades)} trades triggered. "
          f"Applying position sizing...")

    final_trades, ending_capital, curve, taken_ids = apply_position_sizing(all_raw_trades)
    for s in all_signals:
        s["TakenAsTrade"] = s["SignalId"] in taken_ids
    return final_trades, ending_capital, curve, all_signals, rankings


# ------------------------------ EXCEL OUTPUT --------------------------------

TRADE_COLUMNS = ["EntryDate", "EntryTime", "Ticker", "Direction", "Entry", "InitialStop",
                  "Target", "Qty", "RiskAmount", "ExitTime", "ExitPrice", "ExitReason",
                  "Outcome", "PnL", "PnLPct", "CapitalAfter"]
SIGNAL_COLUMNS = ["Date", "DetectedTime", "Ticker", "Direction", "BoxHigh", "BoxLow",
                   "EntryTriggered", "EntryTime", "TakenAsTrade"]


def max_drawdown_pct(curve):
    peak = curve[0]
    max_dd = 0.0
    for value in curve:
        peak = max(peak, value)
        dd = (peak - value) / peak * 100 if peak else 0
        max_dd = max(max_dd, dd)
    return round(max_dd, 2)


def write_excel(trades, starting_capital, ending_capital, capital_curve, all_signals, rankings):
    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"
    ws.append(TRADE_COLUMNS)
    for t in trades:
        ws.append([t.get(c, "") for c in TRADE_COLUMNS])

    wins = sum(1 for t in trades if t["Outcome"] == "WIN")
    total = len(trades)
    win_rate = round(wins / total * 100, 2) if total else 0
    total_pnl = round(sum(t["PnL"] for t in trades), 2)
    return_pct = round((ending_capital - starting_capital) / starting_capital * 100, 2)

    summary_ws = wb.create_sheet("Summary")
    for row in [
        ("Starting Capital", starting_capital), ("Ending Capital", round(ending_capital, 2)),
        ("Total Return %", return_pct), ("Total Trades", total), ("Wins", wins),
        ("Losses", total - wins), ("Win Rate %", win_rate), ("Total P&L", total_pnl),
        ("Max Drawdown %", max_drawdown_pct(capital_curve)),
        ("Risk Per Trade %", RISK_PER_TRADE_PCT), ("Max Daily Risk Cap %", MAX_DAILY_RISK_PCT),
        ("Top N Gainers Scanned Per Day", TOP_N_GAINERS),
        ("Total Wick Rejections Detected", len(all_signals)),
        ("No new entries after", f"{NO_ENTRY_AFTER_HOUR}:{NO_ENTRY_AFTER_MINUTE:02d}"),
        ("Mandatory exit by", f"{EXIT_HOUR}:{EXIT_MINUTE:02d}"),
        ("Max hold candles", MAX_HOLD_CANDLES),
    ]:
        summary_ws.append(row)

    by_symbol = {}
    for t in trades:
        s = by_symbol.setdefault(t["Ticker"], {"trades": 0, "wins": 0, "pnl": 0.0})
        s["trades"] += 1
        s["wins"] += 1 if t["Outcome"] == "WIN" else 0
        s["pnl"] += t["PnL"]
    symbol_ws = wb.create_sheet("BySymbol")
    symbol_ws.append(["Ticker", "Trades", "Wins", "WinRate%", "TotalPnL"])
    for ticker, s in sorted(by_symbol.items(), key=lambda x: -x[1]["pnl"]):
        wr = round(s["wins"] / s["trades"] * 100, 2) if s["trades"] else 0
        symbol_ws.append([ticker, s["trades"], s["wins"], wr, round(s["pnl"], 2)])

    # Separate Bullish / Bearish sheets, each with their own win rate up top
    for direction_label, sheet_name in [("BULLISH", "Bullish"), ("BEARISH", "Bearish")]:
        subset = [t for t in trades if t["Direction"] == direction_label]
        d_ws = wb.create_sheet(sheet_name)
        d_wins = sum(1 for t in subset if t["Outcome"] == "WIN")
        d_total = len(subset)
        d_win_rate = round(d_wins / d_total * 100, 2) if d_total else 0
        d_pnl = round(sum(t["PnL"] for t in subset), 2)
        d_ws.append(["Total Trades", d_total])
        d_ws.append(["Wins", d_wins])
        d_ws.append(["Losses", d_total - d_wins])
        d_ws.append(["Win Rate %", d_win_rate])
        d_ws.append(["Total P&L", d_pnl])
        d_ws.append([])  # blank separator row
        d_ws.append(TRADE_COLUMNS)
        for t in subset:
            d_ws.append([t.get(c, "") for c in TRADE_COLUMNS])

    signals_ws = wb.create_sheet("AllSignals")
    signals_ws.append(SIGNAL_COLUMNS)
    for s in sorted(all_signals, key=lambda x: (x["Date"], x["DetectedTime"])):
        signals_ws.append([s.get(c, "") for c in SIGNAL_COLUMNS])

    rankings_ws = wb.create_sheet("DailyTopGainers")
    rankings_ws.append(["Date", "TopGainerTickers"])
    for day, tickers in sorted(rankings.items()):
        rankings_ws.append([day.strftime("%Y-%m-%d"), ", ".join(tickers)])

    wb.save(OUTPUT_FILE)


# --------------------------- STANDALONE WEB SERVER --------------------------

backtest_status = {"running": False, "started_at": None, "finished_at": None,
                    "result_summary": None, "error": None}


def _run_backtest_background(days):
    backtest_status["running"] = True
    backtest_status["started_at"] = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    backtest_status["error"] = None
    try:
        days = min(days, MAX_LOOKBACK_DAYS)
        trades, ending_capital, curve, all_signals, rankings = run_backtest(days)
        if trades or all_signals:
            write_excel(trades, BACKTEST_CAPITAL, ending_capital, curve, all_signals, rankings)
            wins = sum(1 for t in trades if t["Outcome"] == "WIN")
            wr = round(wins / len(trades) * 100, 1) if trades else 0
            backtest_status["result_summary"] = (
                f"{len(trades)} trades, win rate {wr}%, ended at Rs.{round(ending_capital,2)} "
                f"(started Rs.{BACKTEST_CAPITAL:,.2f}), max drawdown {max_drawdown_pct(curve)}%, "
                f"{len(all_signals)} total wick rejections detected"
            )
        else:
            backtest_status["result_summary"] = "No signals or trades were generated -- check server logs."
    except Exception as e:
        backtest_status["error"] = str(e)
    finally:
        backtest_status["running"] = False
        backtest_status["finished_at"] = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")


class _Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        if self.path.startswith("/run-backtest"):
            self._trigger()
        elif self.path.startswith("/status"):
            self._status()
        elif self.path.startswith("/download"):
            self._download()
        else:
            self._home()

    def do_HEAD(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()

    def _home(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()
        self.wfile.write((
            "Box-Wick (top gainers) backtest service alive.\n\n"
            "Visit /run-backtest to start (optional: ?days=30)\n"
            "Visit /status to check progress.\n"
            "Visit /download to get box_wick_journal.xlsx once finished.\n"
        ).encode("utf-8"))

    def _trigger(self):
        if backtest_status["running"]:
            self.send_response(200)
            self.send_header("Content-Type", "text/plain")
            self.end_headers()
            self.wfile.write(b"A backtest is already running. Check /status.\n")
            return
        query = parse_qs(urlparse(self.path).query)
        days = int(query.get("days", [MAX_LOOKBACK_DAYS])[0])
        threading.Thread(target=_run_backtest_background, args=(days,), daemon=True).start()
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()
        self.wfile.write(
            f"Backtest started ({days} days). Check /status for progress, /download when done.\n".encode("utf-8")
        )

    def _status(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()
        lines = [f"Running: {backtest_status['running']}",
                 f"Started at: {backtest_status['started_at']}",
                 f"Finished at: {backtest_status['finished_at']}",
                 f"Result: {backtest_status['result_summary']}",
                 f"Error (if any): {backtest_status['error']}"]
        self.wfile.write(("\n".join(lines) + "\n").encode("utf-8"))

    def _download(self):
        if not os.path.exists(OUTPUT_FILE):
            self.send_response(404)
            self.send_header("Content-Type", "text/plain")
            self.end_headers()
            self.wfile.write(b"No box_wick_journal.xlsx yet -- visit /run-backtest first.\n")
            return
        with open(OUTPUT_FILE, "rb") as f:
            data = f.read()
        self.send_response(200)
        self.send_header("Content-Type",
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", "attachment; filename=box_wick_journal.xlsx")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, format, *args):
        pass


def start_web_server():
    port = int(os.environ.get("PORT", 8080))
    server = HTTPServer(("0.0.0.0", port), _Handler)
    threading.Thread(target=server.serve_forever, daemon=True).start()
    print(f"Box-Wick backtest server listening on port {port}.")


def main():
    parser = argparse.ArgumentParser(description="Backtest the Box/Wick rejection strategy on previous-day top gainers")
    parser.add_argument("--days", type=int, default=MAX_LOOKBACK_DAYS)
    parser.add_argument("--once", action="store_true",
                         help="Run once immediately and exit (Shell tab / local use). "
                              "Without this flag, starts a web server for Render deployment.")
    args = parser.parse_args()

    if args.once:
        days = min(args.days, MAX_LOOKBACK_DAYS)
        trades, ending_capital, curve, all_signals, rankings = run_backtest(days)
        if not trades and not all_signals:
            print("\nNo signals or trades were generated -- check the [warn] lines above.")
            return
        write_excel(trades, BACKTEST_CAPITAL, ending_capital, curve, all_signals, rankings)
        wins = sum(1 for t in trades if t["Outcome"] == "WIN")
        wr = round(wins / len(trades) * 100, 1) if trades else 0
        print(f"\n{'='*60}\nBACKTEST COMPLETE")
        print(f"  Total wick rejections: {len(all_signals)}")
        print(f"  Total trades:          {len(trades)}")
        print(f"  Win rate:              {wr}%")
        print(f"  Starting capital:      Rs.{BACKTEST_CAPITAL:,.2f}")
        print(f"  Ending capital:        Rs.{ending_capital:,.2f}")
        print(f"  Max drawdown:          {max_drawdown_pct(curve)}%")
        print(f"  Full detail in:        {OUTPUT_FILE}\n{'='*60}")
        return

    start_web_server()
    print("Waiting for a backtest to be triggered via /run-backtest. Press Ctrl+C to stop.")
    try:
        while True:
            time.sleep(60)
    except KeyboardInterrupt:
        print("\nStopped by user. Goodbye!")
        sys.exit(0)


if __name__ == "__main__":
    main()
